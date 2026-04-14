"""
06_export_powerbi.py
────────────────────
WHAT THIS FILE DOES:
  - Creates the final Power BI-ready Excel workbook with all sheets pre-formatted
  - Each sheet is designed for a specific Power BI visual type
  - Also creates a "Power BI Connection Guide" sheet with exact steps

WHY A DEDICATED EXPORT:
  Power BI needs data in a very specific shape:
  - One row per observation (no merged cells, no header gaps)
  - Date column must be clean datetime, not text
  - Device names must be in rows, NOT as column headers (unpivoted format)
    for pie/donut/bar charts that need a legend field

SHEETS IN OUTPUT EXCEL:
  Daily_Master         → wide format (1 row per day, devices as columns)
  Daily_Unpivoted      → long format (1 row per device per day) — for device charts
  Forecast_7Day        → 7-day forecast with upper/lower bands
  Summary_KPIs         → flat table of all KPI values — for Card visuals
  AC_Savings_Detail    → granular AC savings data — for AC Deep Dive page
  Lights_Detail        → granular lighting data — for Lights page
  PowerBI_Guide        → step-by-step connection and setup instructions

HOW TO RUN:
  python src/06_export_powerbi.py
  (Run AFTER 05_charts.py — all previous stages must have run)
"""

import sys
import pickle
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import PROCESSED_DIR, OUTPUTS_DIR


def build_unpivoted(master: pd.DataFrame) -> pd.DataFrame:
    """
    Converts wide format (devices as columns) to long format (devices as rows).

    Wide format — ONE row per day:
      date     | sockets | front_ac | back_ac | server
      08-Dec   |  47.2   |  36.4    |  36.4   |  6.7

    Long format — FOUR rows per day (one per device):
      date     | device    | consumption_kwh
      08-Dec   | Sockets   |  47.2
      08-Dec   | Front AC  |  36.4
      08-Dec   | Back AC   |  36.4
      08-Dec   | Server Rm |   6.7

    WHY: Power BI's pie, donut, and stacked charts need the 'device' as
    a FIELD, not as 10 separate column names. Long format makes this possible.
    """
    device_map = {
        "sockets":   "Sockets",
        "front_ac":  "Front AC",
        "back_ac":   "Back AC",
        "server":    "Server Room",
        "total_lights": "All Lights",
    }

    rows = []
    for _, row in master.iterrows():
        for col, label in device_map.items():
            rows.append({
                "date":             row["date"],
                "day_name":         row["day_name"],
                "is_weekend":       row["is_weekend"],
                "device":           label,
                "consumption_kwh":  round(row[col], 4),
                "is_active":        row["is_active"],
            })

    return pd.DataFrame(rows).sort_values(["date", "device"]).reset_index(drop=True)


def build_kpi_table(master: pd.DataFrame, forecast_pkg: dict) -> pd.DataFrame:
    """
    Flat table of KPI values — exactly what Power BI Card visuals need.
    One row = one KPI. The 'Value_Numeric' column is used for the card.
    The 'Display_Value' column is pre-formatted for text labels.
    """
    active = master[master["is_active"]]

    total_kwh    = master["total"].sum()
    total_saved  = master["total_ac_saved"].sum()
    baseline     = master["total_ac_baseline"].sum()
    savings_rate = total_saved / baseline * 100 if baseline > 0 else 0
    co2_saved    = master["carbon_saved_kg"].sum()
    co2_emitted  = master["carbon_emitted_kg"].sum()
    fc_range     = forecast_pkg["forecast_df"]["forecast_kwh"]

    kpis = [
        ("Total Consumption",    "kWh",  round(total_kwh, 1),   f"{total_kwh:.1f} kWh"),
        ("Total Energy Saved",   "kWh",  round(total_saved, 1), f"{total_saved:.1f} kWh"),
        ("Overall Savings Rate", "%",    round(savings_rate, 1),f"{savings_rate:.1f}%"),
        ("CO2 Emitted",          "kg",   round(co2_emitted, 1), f"{co2_emitted:.1f} kg CO₂"),
        ("CO2 Saved",            "kg",   round(co2_saved, 1),   f"{co2_saved:.1f} kg CO₂"),
        ("Active Days",          "days", int(master["is_active"].sum()), str(int(master["is_active"].sum()))),
        ("Weekend/Off Days",     "days", int((~master["is_active"]).sum()), str(int((~master["is_active"]).sum()))),
        ("Peak Consumption",     "kWh",  round(master["total"].max(), 1), f"{master['total'].max():.1f} kWh"),
        ("Peak Day",             "date", None, master.loc[master["total"].idxmax(), "date"].strftime("%d %b")),
        ("Avg Daily (Active)",   "kWh",  round(active["total"].mean(), 1), f"{active['total'].mean():.1f} kWh"),
        ("Forecast Min (7-day)", "kWh",  round(fc_range.min(), 1), f"{fc_range.min():.1f} kWh"),
        ("Forecast Max (7-day)", "kWh",  round(fc_range.max(), 1), f"{fc_range.max():.1f} kWh"),
        ("Front AC Total",       "kWh",  round(master["front_ac"].sum(), 1), f"{master['front_ac'].sum():.1f} kWh"),
        ("Back AC Total",        "kWh",  round(master["back_ac"].sum(), 1), f"{master['back_ac'].sum():.1f} kWh"),
        ("All Lights Total",     "kWh",  round(master["total_lights"].sum(), 1), f"{master['total_lights'].sum():.1f} kWh"),
    ]

    return pd.DataFrame(kpis, columns=["KPI_Name", "Unit", "Value_Numeric", "Display_Value"])


def build_ac_detail(master: pd.DataFrame) -> pd.DataFrame:
    """
    Granular AC table — used for the Power BI AC Deep Dive page.
    Contains both Front and Back AC in long format with savings columns.
    """
    rows = []
    for _, row in master.iterrows():
        for unit, used_col, saved_col, base_col in [
            ("Front AC", "front_ac", "front_ac_saved", "front_ac_baseline"),
            ("Back AC",  "back_ac",  "back_ac_saved",  "back_ac_baseline"),
        ]:
            used     = row[used_col]  if pd.notna(row[used_col])  else 0
            saved    = row[saved_col] if pd.notna(row[saved_col]) else 0
            baseline = row[base_col]  if pd.notna(row[base_col])  else 0
            rate     = saved / baseline * 100 if baseline > 0 else 0

            rows.append({
                "date":         row["date"],
                "day_name":     row["day_name"],
                "is_weekend":   row["is_weekend"],
                "ac_unit":      unit,
                "used_kwh":     round(used, 4),
                "saved_kwh":    round(saved, 4),
                "baseline_kwh": round(baseline, 4),
                "savings_pct":  round(rate, 2),
                "is_active":    row["is_active"],
            })

    return pd.DataFrame(rows).sort_values(["date", "ac_unit"]).reset_index(drop=True)


def build_lights_detail(master: pd.DataFrame) -> pd.DataFrame:
    """
    Granular lights table — used for the Power BI Lights Analysis page.
    """
    zone_map = {
        "lib_front_kwh": "Front Area Lights",
        "lib_hall_kwh":  "Hall / Lobby Lights",
        "lib_back_kwh":  "Back Area Lights",
    }

    rows = []
    for _, row in master.iterrows():
        for col, label in zone_map.items():
            kwh = row[col] if pd.notna(row[col]) else 0
            rows.append({
                "date":       row["date"],
                "day_name":   row["day_name"],
                "is_weekend": row["is_weekend"],
                "zone":       label,
                "kwh":        round(kwh, 5),
                "is_active":  kwh > 0.01,
            })

    return pd.DataFrame(rows).sort_values(["date", "zone"]).reset_index(drop=True)


def build_guide() -> pd.DataFrame:
    """
    Step-by-step guide for connecting Power BI to this Excel file.
    """
    steps = [
        ("STEP 1", "Install Power BI Desktop",
         "Download free from powerbi.microsoft.com/desktop. Use Desktop, NOT Power BI Service (online)."),
        ("STEP 2", "Open Power BI and Connect to Excel",
         "Home → Get Data → Excel Workbook → browse to BuildINT_PowerBI_Data.xlsx → select ALL sheets → Load"),
        ("STEP 3", "Open Power Query Editor",
         "Home → Transform Data. For each table: verify Date column = Date type, number columns = Decimal Number. Close & Apply."),
        ("STEP 4", "Create Relationships (if needed)",
         "Model view → drag 'date' from Daily_Master to Forecast_7Day.date. This links tables so date slicers affect all visuals."),
        ("STEP 5 — Card Visuals (KPI numbers)",
         "Visualizations → Card → drag 'Value_Numeric' from Summary_KPIs → filter by KPI_Name using a filter pane (not a slicer)."),
        ("STEP 6 — Stacked Bar (Daily Consumption)",
         "Visualization → Stacked bar chart → X-axis: date (Daily_Master) → Values: sockets, front_ac, back_ac, server"),
        ("STEP 7 — Donut (Device Share)",
         "Visualization → Donut chart → Legend: device (Daily_Unpivoted) → Values: consumption_kwh"),
        ("STEP 8 — Line Chart (Trend)",
         "Visualization → Line chart → X-axis: date → Values: consumption_kwh (Daily_Master: total)"),
        ("STEP 9 — Forecast Chart",
         "Line chart → X-axis: date (mix of Historical + Forecast_7Day) → Values: actual_kwh, forecast_kwh → add upper/lower as shaded area"),
        ("STEP 10 — Date Slicer",
         "Visualization → Slicer → Field: date → Format → Slicer settings → Between. Place at top."),
        ("STEP 11 — AC Deep Dive Page",
         "Add new page. Line/bar chart from AC_Savings_Detail: X-axis: date, Legend: ac_unit, Values: used_kwh, saved_kwh"),
        ("STEP 12 — Lights Analysis Page",
         "Add new page. Line chart from Lights_Detail: X-axis: date, Legend: zone, Values: kwh"),
        ("STEP 13 — Export to PDF",
         "File → Export → Export to PDF. This captures all pages for sharing."),
        ("TIP — DAX Savings Rate Measure",
         "Right-click table → New Measure → type:  Savings Rate = DIVIDE(SUM(AC_Savings_Detail[saved_kwh]), SUM(AC_Savings_Detail[baseline_kwh])) * 100"),
        ("TIP — Weekend Filter",
         "Add Slicer on is_weekend (True/False) from Daily_Master to let users toggle weekday vs weekend view instantly."),
    ]

    return pd.DataFrame(steps, columns=["Step", "Action", "Details"])


def style_excel(wb):
    """
    Applies professional styling to the Excel workbook.
    - Navy header rows, white bold text
    - Alternating row colors for readability
    - Auto-fit column widths
    """
    NAVY   = "1F3864"
    LIGHT  = "EBF3FB"
    WHITE  = "FFFFFF"

    header_font  = Font(bold=True, color=WHITE, size=11)
    header_fill  = PatternFill("solid", fgColor=NAVY)
    alt_fill     = PatternFill("solid", fgColor=LIGHT)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left", vertical="center", wrap_text=False)

    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        # Style header row
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
            cell.border    = border

        # Style data rows
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = alt_fill if i % 2 == 0 else None
            for cell in row:
                if fill:
                    cell.fill = fill
                cell.alignment = left_align
                cell.border    = border

        # Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        # Freeze top row
        ws.freeze_panes = "A2"

    return wb


def main():
    print("=" * 60)
    print("  STAGE 6: POWER BI EXPORT")
    print("=" * 60)

    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)

    master = pd.read_pickle(PROCESSED_DIR / "master.pkl")

    with open(PROCESSED_DIR / "forecast.pkl", "rb") as f:
        forecast_pkg = pickle.load(f)

    # Build all output tables
    print("  Building tables...")
    unpivoted    = build_unpivoted(master)
    kpi_table    = build_kpi_table(master, forecast_pkg)
    ac_detail    = build_ac_detail(master)
    lights_detail = build_lights_detail(master)
    guide        = build_guide()
    forecast_df  = forecast_pkg["forecast_df"]
    historical_df = forecast_pkg["historical_df"]

    print(f"    Daily_Master:      {master.shape}")
    print(f"    Daily_Unpivoted:   {unpivoted.shape}")
    print(f"    Forecast_7Day:     {forecast_df.shape}")
    print(f"    Summary_KPIs:      {kpi_table.shape}")
    print(f"    AC_Savings_Detail: {ac_detail.shape}")
    print(f"    Lights_Detail:     {lights_detail.shape}")

    # Write to Excel
    output_path = OUTPUTS_DIR / "BuildINT_PowerBI_Data.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        master.to_excel(       writer, sheet_name="Daily_Master",      index=False)
        unpivoted.to_excel(    writer, sheet_name="Daily_Unpivoted",   index=False)
        forecast_df.to_excel(  writer, sheet_name="Forecast_7Day",     index=False)
        historical_df.to_excel(writer, sheet_name="Historical",        index=False)
        kpi_table.to_excel(    writer, sheet_name="Summary_KPIs",      index=False)
        ac_detail.to_excel(    writer, sheet_name="AC_Savings_Detail", index=False)
        lights_detail.to_excel(writer, sheet_name="Lights_Detail",     index=False)
        guide.to_excel(        writer, sheet_name="PowerBI_Guide",     index=False)

        # Apply styling
        wb = writer.book
        wb = style_excel(wb)

    print(f"\n  ✅ Exported: {output_path}")
    print(f"     Open this file in Power BI: Home → Get Data → Excel Workbook")
    print(f"     See the 'PowerBI_Guide' sheet for exact step-by-step instructions.")

    print("\n" + "=" * 60)
    print("  ✅ Stage 6 Complete. Full pipeline done!")
    print("=" * 60)
    print()
    print("  DELIVERABLES:")
    print(f"    📊 Power BI Data:  data/outputs/BuildINT_PowerBI_Data.xlsx")
    print(f"    📈 Charts (PNGs):  reports/charts/  (8 charts)")
    print(f"    📦 Analysis data:  data/processed/  (pkl files)")
    print()
    print("  NEXT STEPS:")
    print("    1. Open BuildINT_PowerBI_Data.xlsx in Power BI Desktop")
    print("    2. Follow the PowerBI_Guide sheet instructions")
    print("    3. Export Power BI to PDF → paste screenshots into PPT")


if __name__ == "__main__":
    main()